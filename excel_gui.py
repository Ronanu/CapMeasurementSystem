import tkinter as tk
from tksheet import Sheet
from openpyxl import load_workbook

class MessGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Messparameter Editor")
        self.file_path = "Messparameter.xlsx"

        # Fenstergröße maximieren und skalierbar machen
        self.root.geometry("800x800")
        #self.root.grid_rowconfigure(0, weight=1)  # Tabelle wächst mit dem Fenster
        # self.root.grid_columnconfigure(0, weight=1)

        # Hauptframe für alles
        self.main_frame = tk.Frame(root)
        self.main_frame.grid(row=0, column=0, sticky="nsew")

        # Tabelle mit tksheet
        self.sheet = Sheet(self.main_frame)
        self.sheet.grid(row=0, column=0, sticky="nsew")

        # Buttons unten
        self.button_frame = tk.Frame(self.main_frame)
        self.button_frame.grid(row=1, column=0, sticky="ew", pady=5)

        self.start_btn = tk.Button(self.button_frame, text="Start Messung", command=self.save_parameters)
        self.start_btn.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

        self.store_btn = tk.Button(self.button_frame, text="Store Parameters", command=self.save_parameters)
        self.store_btn.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10)

        self.load_excel()

    def load_excel(self):
        """Lädt die Excel-Datei und stellt sicher, dass nur die Werte-Spalte editierbar ist"""
        try:
            self.wb = load_workbook(self.file_path)  # Excel-Datei laden
            self.ws = self.wb.active

            # Spaltennamen aus der ersten Zeile
            headers = [cell.value for cell in self.ws[1]]

            # Daten aus Excel-Tabelle extrahieren
            self.data = [list(row) for row in self.ws.iter_rows(min_row=2, values_only=True)]

            # Daten in die Tabelle setzen
            self.sheet.set_sheet_data(self.data)
            self.sheet.headers(headers)

            # Nur die Werte-Spalte (Index 1) bleibt editierbar
            readonly_cols = [0, 2]  # "Parameter" (0) & "Einheit" (2) gesperrt
            self.sheet.readonly_columns(readonly_cols)

            # Automatische Spaltenbreite
            self.sheet.enable_bindings("all")

        except FileNotFoundError:
            print(f"Fehler: '{self.file_path}' nicht gefunden!")

    def save_parameters(self):
        """Speichert nur die Werte-Spalte zurück in die bestehende Datei, ohne Formatierung zu verlieren"""
        # **🔹 Automatische Bestätigung der letzten Eingabe (Enter simulieren)**
        active_cell = self.sheet.get_currently_selected()
        if isinstance(active_cell, tuple) and len(active_cell) >= 2:
            row, col = active_cell[:2]
            if col == 1:  # Nur falls in der Werte-Spalte
                new_value = self.sheet.get_cell_data(row, col)  # Den aktuellen Wert holen
                self.sheet.set_cell_data(row, col, new_value)  # Wert explizit setzen, um ihn zu speichern

        self.sheet.focus_set()  # Fokus auf das Hauptfenster setzen (Enter-Simulation)
        self.sheet.deselect()   # Alle Markierungen entfernen, um sicherzugehen

        new_data = self.sheet.get_sheet_data()  # Holt die aktuellen Werte aus der Tabelle

        # Werte-Spalte aktualisieren (Index 1 ist editierbar)
        for i, row in enumerate(new_data, start=2):  # Startet ab Zeile 2 (erste Zeile ist Header)
            if len(row) > 1:  # Sicherstellen, dass es eine Werte-Spalte gibt
                self.ws.cell(row=i, column=2, value=row[1])  # Nur Spalte "Wert" aktualisieren

        # Speichern der Datei mit den neuen Werten
        self.wb.save(self.file_path)
        self.wb.close()
        print(f"Neue Messparameter gespeichert in '{self.file_path}'.")

if __name__ == "__main__":
    root = tk.Tk()
    app = MessGUI(root)
    root.mainloop()
